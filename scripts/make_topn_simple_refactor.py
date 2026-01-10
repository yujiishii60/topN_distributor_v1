# scripts/make_topn_simple_refactor.py
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import math
from openpyxl.formatting.rule import Rule
from calendar import monthrange
import json
import re
from datetime import datetime
from string import Template

CATEGORY_MAP = {
    "1": "寿司", "2": "米飯", "3": "温惣菜",
    "4": "冷総菜", "5": "軽食", "6": "魚惣菜",
}

def store_folder_name(store_id) -> str:
    """店番フォルダ名（2桁ゼロ埋め）"""
    s = str(store_id).strip()
    try:
        return f"{int(s):02d}"
    except Exception:
        # 数字以外はそのまま（必要ならここでsanitize）
        return s

def _month_keys_from_dates(dates):
    """dates(list[str or date]) → {'2024-12', '2025-01'} のような集合"""
    dt = pd.to_datetime(dates).date
    return {f"{d.year:04d}-{d.month:02d}" for d in dt}

def copy_conditional_formatting(ws_dst, ws_src):
    """
    copy_worksheet で失われがちな条件付き書式を、TEMPLATE から新シートへ再適用する。
    レイアウトが同一（セル座標が同じ）前提で、そのまま同レンジへ貼る。
    """
    cf_src = ws_src.conditional_formatting
    # openpyxlの内部構造はバージョンで差があります。代表的な2系統に対応。
    if hasattr(cf_src, "cf_rules"):  # 3.1系で公開属性がある場合
        items = cf_src.cf_rules.items()
    else:  # 旧来: _cf_rules にレンジ→ルールlist が入っていることが多い
        items = getattr(cf_src, "_cf_rules", {}).items()

    for rng, rules in items:
        for rule in rules:
            # そのまま add すると同じオブジェクト参照になることがあるので clone 相当を作る
            new_rule = Rule(
                type=rule.type,
                dxf=rule.dxf,
                formula=list(rule.formula) if hasattr(rule, "formula") else None,
                operator=getattr(rule, "operator", None),
                text=getattr(rule, "text", None),
                timePeriod=getattr(rule, "timePeriod", None),
                rank=getattr(rule, "rank", None),
                percent=getattr(rule, "percent", None),
                stopIfTrue=getattr(rule, "stopIfTrue", False),
            )
            # カラースケール/データバーなど複合型も転写
            for attr in ("colorScale", "dataBar", "iconSet"):
                if hasattr(rule, attr) and getattr(rule, attr):
                    setattr(new_rule, attr, getattr(rule, attr))
            ws_dst.conditional_formatting.add(rng, new_rule)

# === CSV読込 ===
def load_sales(root: Path, dates=None) -> pd.DataFrame:
    """
    data/material/YYYY/IT_YYYYMM.csv を必要分だけ読む（年月またぎ対応）。
    返り値は標準列:
      date (datetime.date), store_id (str), category_large (str),
      jan (str), name (str), amount (float), qty (float), discount (float)
    同一 (date, store, category_large, jan) は合算（nameはfirst）
    """
    root = Path(root)
    # 読むべき年月を決定
    if dates:
        dates = [pd.to_datetime(d).date() for d in dates]
        ym_keys = sorted({(d.year, d.month) for d in dates})
    else:
        # dates未指定ならルート配下を総当り（従来動作）
        ym_keys = []
        for y_dir in (root.glob("*")):
            if y_dir.is_dir() and y_dir.name.isdigit():
                for f in y_dir.glob("IT_*.csv"):
                    # IT_YYYYMM.csv から (YYYY,MM) を推定
                    stem = f.stem  # IT_202501
                    y = int(stem.split("_")[1][:4])
                    m = int(stem.split("_")[1][4:6])
                    ym_keys.append((y, m))
        ym_keys = sorted(set(ym_keys))

    files = []
    for y, m in ym_keys:
        f = root / f"{y}" / f"IT_{y}{m:02d}.csv"
        if f.exists():
            files.append(f)

    if not files:
        raise FileNotFoundError(f"no monthly files for {ym_keys} under {root}")

    # 読み込み＋列標準化
    def _read_any(p: Path) -> pd.DataFrame:
        for enc in ("cp932", "utf-8-sig", "utf-8"):
            try:
                return pd.read_csv(p, encoding=enc)
            except Exception:
                continue
        return pd.read_csv(p)  # 最後の保険

    rename_map = {
        "売上日": "date",
        "店舗コード": "store_id",
        "大分類コード": "category_large",
        "中分類コード": "category_middle",
        "小分類コード": "category_small",
        "JANコード": "jan",
        "品名漢字": "name",
        "総売上金額": "amount",
        "総売上数量": "qty",
        "値引金額": "discount",
    }

    df = pd.concat((_read_any(f) for f in files), ignore_index=True)
    df = df.rename(columns=rename_map)

    # 型正規化
    df["date"] = pd.to_datetime(df["date"], errors="coerce").dt.date
    df["store_id"] = df["store_id"].astype(str)
    df["category_large"] = df["category_large"].astype(str)
    df["jan"] = df["jan"].astype(str)
    for col in ("amount", "qty", "discount"):
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0.0).astype(float)
        else:
            df[col] = 0.0

    if dates:
        use = set(dates)
        df = df[df["date"].isin(use)]

    # 4倍問題の再発防止：category_large を含めて集約（←ここが重要）
    agg_map = {"amount": "sum", "qty": "sum", "discount": "sum", "name": "first"}
    df = (df.groupby(["date", "store_id", "category_large", "jan"], as_index=False)
            .agg(agg_map))

    return df

# === 店舗マスター ===
# === store_master 読み込み（store/name/short_name 想定） ===
def load_store_master(path: Path) -> dict[str, str]:
    import pandas as pd
    sm = pd.read_excel(path)

    # 列名を内部統一
    sm = sm.rename(columns={
        "store": "store_id",
        "name": "store_name",
        "short_name": "short_name",
    })

    # 型そろえ（店番は文字列化）
    sm["store_id"] = sm["store_id"].astype(str)

    # short_name が欠けてたら store_name で補完
    if "short_name" not in sm.columns:
        sm["short_name"] = sm["store_name"]
    sm["short_name"] = sm["short_name"].fillna(sm["store_name"])

    # 辞書 { "1": "神栖店", ... } を返す
    return dict(zip(sm["store_id"], sm["short_name"]))

# === 大分類・日付でフィルタ ===
def filter_sales(df: pd.DataFrame, category: str, dates: list[str]) -> pd.DataFrame:
    df = df[df["category_large"].astype(str) == str(category)]
    dates = [pd.to_datetime(d).date() for d in dates]
    df = df[df["date"].isin(dates)]
    return df

# === 店舗×日付TopN抽出 ===
def build_topn(df: pd.DataFrame, top_n=30) -> dict:
    result = {}
    for (store, date), g in df.groupby(["store_id", "date"]):
        g = g.sort_values("amount", ascending=False).head(top_n)
        total = g["amount"].sum()
        g["構成比"] = g["amount"] / total if total else 0
        result.setdefault(store, {})[date] = g
    return result

from openpyxl import Workbook

def _add_pages_for_one_store(wb, ws_tpl, store, store_short_name, dates, day_map, cat_name, event_name,
                             total_all_dict, total_cat_dict, category, make_title):
    block_offsets = [0, 8, 16, 24]
    total_days = len(dates)
    num_pages = math.ceil(total_days / 4)

    for page in range(num_pages):
        ws = wb.copy_worksheet(ws_tpl)
        copy_conditional_formatting(ws, ws_tpl)
        ws.title = f"{store}({page+1})"

        page_dates = dates[page*4 : (page+1)*4]
        # 代表日とページ番号（タイトル用）
        page_no = page + 1
        if page_dates:
            date_str = pd.to_datetime(page_dates[0]).strftime("%Y-%m-%d")
        else:
            date_str = ""
        # タイトル（A1）
        ws["A1"].value = make_title(date_str, page_no)

        for block_idx, d in enumerate(page_dates):
            if block_idx >= 4: break
            col_offset = block_offsets[block_idx]
            d_date = pd.to_datetime(d).date()
            df_day = day_map.get(d_date)
            if df_day is None or df_day.empty:
                continue

            # ブロックヘッダ
            year2 = str(pd.to_datetime(d).year)[2:]
            mmdd  = pd.to_datetime(d).strftime("%m/%d")
            ws.cell(row=2, column=1+col_offset, value=year2)
            ws.cell(row=2, column=2+col_offset, value=mmdd)
            ws.cell(row=2, column=3+col_offset, value=store_short_name)
            ws.cell(row=2, column=4+col_offset, value=f"{cat_name}単品")

            # 見出し
            headers = ["順位", "商品名", "売上金額", "売上数量", "値引金額", "値引率"]
            for i, h in enumerate(headers):
                ws.cell(row=3, column=1+col_offset+i, value=h)

            # 明細（df_day は降順TopN想定）
            for rank, row in enumerate(df_day.itertuples(index=False), start=1):
                if rank > 35: break
                r = 3 + rank
                amt  = float(getattr(row, "amount", 0) or 0.0)
                qty  = float(getattr(row, "qty", 0) or 0.0)
                disc = float(getattr(row, "discount", 0) or 0.0)
                rate = (disc/amt) if amt else 0.0

                ws.cell(r, 1+col_offset, rank)
                ws.cell(r, 2+col_offset, getattr(row, "name", ""))
                ws.cell(r, 3+col_offset, amt)
                ws.cell(r, 4+col_offset, qty)
                ws.cell(r, 5+col_offset, disc)
                c = ws.cell(r, 6+col_offset, rate)
                c.number_format = "0.00%"

            # フッタ
            row_base = 39
            ws.cell(row=row_base+1, column=1+col_offset, value="惣菜売上金額")
            ws.cell(row=row_base+2, column=1+col_offset, value=f"{cat_name}売上金額")
            ws.cell(row=row_base+3, column=1+col_offset, value=f"{cat_name}構成比")

            total_store_amount = total_all_dict.get((d_date, store), 0.0)
            total_cat_amount   = total_cat_dict.get((d_date, store), 0.0)
            ratio = (total_cat_amount/total_store_amount) if total_store_amount else 0.0

            ws.cell(row=row_base+1, column=3+col_offset, value=total_store_amount)
            ws.cell(row=row_base+2, column=3+col_offset, value=total_cat_amount)
            ws.cell(row=row_base+3, column=3+col_offset, value=ratio)
            ws.cell(row=row_base+3, column=3+col_offset).number_format = "0.00%"

def save_per_store_files(master_path: Path, out_root: Path, category_name: str):
    """
    生成済みのマスターExcel(master_path)を基に、
    店番ごとのシートだけ残して新規ファイルとして保存する。
    フォーマット・条件付き書式・書式は維持される。
    """
    out_root = Path(out_root)
    out_root.mkdir(parents=True, exist_ok=True)

    # 一度開いてシート→店番の対応を取得
    wb_probe = load_workbook(master_path)
    store_to_sheets: dict[str, list[str]] = {}
    for s in wb_probe.sheetnames:
        m = re.match(r"^(\d+)\((\d+)\)$", s)  # 例: "25(1)"
        if m:
            sid = m.group(1)
            store_to_sheets.setdefault(sid, []).append(s)
    wb_probe.close()

    # 店番ごとにファイル生成：毎回マスターを開きなおして不要シートを削除
    for sid, sheets in store_to_sheets.items():
        wb = load_workbook(master_path)
        for name in list(wb.sheetnames):
            if name not in sheets:
                ws = wb[name]
                wb.remove(ws)

        # 保存先: split/<店番>/<店番>_<大分類名>単品データ.xlsx
        sid2 = store_folder_name(sid)
        out_dir = out_root / sid2
        out_dir.mkdir(parents=True, exist_ok=True)
        out_path = out_dir / f"{sid2}_{category_name}単品データ.xlsx"
        wb.save(out_path)
        wb.close()

# === Excel書き出し ===
def write_excel(template_path, out_path, topn_dict, store_names, category, dates,
                event_name, df_sales_all=None, split_by_store=False, split_dir="",
                title_template="{event} {date} {cat}単品データ ({page})",
                no_date_in_title=False):
    
    def _dates_to_range(dates: list[str]) -> str:
        # ['2024-12-24','2025-01-03'] → '2024-12–2025-01'（同月なら '2024-12'）
        try:
            ds = sorted(datetime.strptime(str(d), "%Y-%m-%d") for d in dates)
        except Exception:
            return ",".join(map(str, dates))
        if not ds: 
            return ""
        a, b = ds[0], ds[-1]
        return f"{a.year}-{a.month:02d}" if (a.year==b.year and a.month==b.month) else f"{a.year}-{a.month:02d}–{b.year}-{b.month:02d}"

    def _build_title_from_template(event_name: str | None,
                                  category_code: int | str,
                                  dates_list: list[str],
                                  page_no: int,
                                  tmpl: str | None,
                                  cat_name_resolver) -> str:
        """
        イベント名が空→テンプレ採用。テンプレ空→既定テンプレ。
        イベント名あり→従来の「{event} {cat}単品データ（{page}）」優先。
        """
        cat = cat_name_resolver(category_code)
        ev = (event_name or "").strip()

        # イベント名が入っていれば従来優先
        if ev:
            return f"{ev} {cat}単品データ（{page_no}）"

        # イベント名が空→テンプレ（未指定なら既定テンプレ）
        default_tmpl = "{yy}年 {range} {cat}単品データ（{page}）"
        tmpl = (tmpl or default_tmpl).strip()

        # === 日付の範囲から代表日を決める ===
        if dates_list:
            ds_sorted = sorted(dates_list)
            first_date_str = str(ds_sorted[0])
            last_date_str  = str(ds_sorted[-1])  # ★追加：末尾日
        else:
            first_date_str = last_date_str = ""

        first_date_short = first_date_str[5:].replace("-", "/") if first_date_str else ""
        last_date_short  = last_date_str[5:].replace("-", "/") if last_date_str else ""

        values = {
            "event": ev,
            "cat": cat,
            "category": str(category_code),
            "dates": ",".join(map(str, dates_list)),
            "dates_short": ",".join(str(d)[5:].replace("-", "/") for d in dates_list),
            "range": _dates_to_range([str(d) for d in dates_list]),
            "year": (last_date_str[:4] if last_date_str else ""),   # ★末尾日で決定
            "yy":   (last_date_str[2:4] if last_date_str else ""),  # ★末尾日で決定
            "date": last_date_str,
            "date_short": last_date_short,
            "page": str(page_no),
        }


        # 1) $var 形式の置換
        s = Template(tmpl).safe_substitute(values)
        # 2) {var} 形式の置換（1で未展開の {} をここで仕上げ）
        try:
            return s.format(**values)
        except Exception:
            return s  # それでも失敗したら、展開できた分だけ返す


    # --- タイトル生成のユーティリティ（外部マップ対応） ---
    def _cat_name_from_code(code: int | str) -> str:
        cfg = Path("config/category_map.json")
        if cfg.exists():
            try:
                m = json.loads(cfg.read_text(encoding="utf-8"))
                return m.get(str(code), str(code))
            except Exception:
                pass
        builtin = {"1":"寿司","2":"弁当","3":"温総菜","4":"冷総菜","5":"軽食","6":"魚惣菜"}
        return builtin.get(str(code), str(code))

    # --- タイトルテンプレ & 関数（ここだけ1回定義） ---
    eff_tmpl = "{event} {cat}単品データ ({page})" if no_date_in_title else title_template
    cat_name = _cat_name_from_code(category)
    ev = (event_name or "").strip() or "（無題）"
    
    # --- タイトル関数（イベント名が空ならテンプレ発動） ---
    def make_title(date_str: str, page_no: int) -> str:
        # no_date_in_title=True の時はテンプレに日付情報を渡さない
        dlist = [] if no_date_in_title else [str(d) for d in dates]
        return _build_title_from_template(
            event_name=event_name,
            category_code=category,
            dates_list=dlist,
            page_no=page_no,
            tmpl=title_template,
            cat_name_resolver=_cat_name_from_code,
        )


    wb0 = load_workbook(template_path)
    ws_tpl0 = wb0["TEMPLATE"]
    cat_name = CATEGORY_MAP.get(str(category), str(category))

    # === 合計のための辞書（全惣菜 / 大分類）を先に作る ===
    # df_sales_all は load_sales 後のもの（同一キー集約済み）なので、
    # 「大分類合計」は dates と store で再集計が必要。
    # → 全体合計はそのまま、カテゴリ合計は別途 df を再作成して辞書化。
    g_all = df_sales_all.copy()
    # 使う日だけに絞る
    use_dates = set(pd.to_datetime(dates).date)
    g_all = g_all[g_all["date"].isin(use_dates)]

    # 全惣菜（全カテゴリ）
    total_all_dict = g_all.groupby(["date", "store_id"])["amount"].sum().to_dict()

    # 大分類（category）合計：元CSVを読む時点で category_large を落としているため、
    # ここでは元データ（未集約）から計算するのが理想だが、
    # 今回は aggregate_topn に渡した df_sales（未フィルタ）を別に保持している想定がないので、
    # df_sales_all に category_large が無いケースを考慮し、呼び出し側で
    # 「df_sales_all は category_large を含む DataFrame」を渡す方針で運用する。
    # もし含まない場合は、上位呼び出しで別に df_raw を渡す実装拡張が必要。
    if "category_large" in df_sales_all.columns:
        g_cat = df_sales_all.copy()
        g_cat = g_cat[g_cat["date"].isin(use_dates)]
        g_cat = g_cat[g_cat["category_large"].astype(str) == str(category)]
        total_cat_dict = g_cat.groupby(["date", "store_id"])["amount"].sum().to_dict()
    else:
        # 最低限のフォールバック（TopNの金額合計を使用）
        total_cat_dict = {}
        for store, day_map in topn_dict.items():
            for d, df_day in day_map.items():
                total_cat_dict[(d, store)] = float(df_day["amount"].sum())

    # === 出力先（店別）ルート
    if split_by_store:
        # ← ここは out_store_dir ではなく split_dir に統一
        base_dir = Path(split_dir) if split_dir else Path(out_path).parent / "stores"
        base_dir.mkdir(parents=True, exist_ok=True)

        # 店ごとにテンプレから新規WBを作り、該当店のシートだけ収めて保存
        for store in sorted(topn_dict.keys(), key=lambda x: int(x)):
            wb = load_workbook(template_path)
            ws_tpl = wb["TEMPLATE"]

            _add_pages_for_one_store(
                wb, ws_tpl,
                store=store,
                store_short_name=store_names.get(store, ""),
                dates=dates,
                day_map=topn_dict[store],
                cat_name=cat_name,
                event_name=event_name,
                total_all_dict=total_all_dict,
                total_cat_dict=total_cat_dict,
                category=category,
                make_title=make_title,                
            )

            # テンプレシートが残っていれば削除（存在チェック）
            if "TEMPLATE" in wb.sheetnames:
                del wb["TEMPLATE"]

            # 1番フォルダ / "1_冷総菜単品データ.xlsx"
            sid2 = store_folder_name(store)
            subdir = base_dir / sid2
            subdir.mkdir(parents=True, exist_ok=True)
            safe_cat = f"{cat_name}".replace("/", "／").replace("\\", "／")
            out_file = subdir / f"{sid2}_{safe_cat}単品データ.xlsx"
            wb.save(out_file)

        print(f"[ok] split saved → {base_dir}")

    else:
        # 既存：全店を1冊に
        wb = load_workbook(template_path)
        ws_tpl = wb["TEMPLATE"]

        for store in sorted(topn_dict.keys(), key=lambda x: int(x)):
            _add_pages_for_one_store(
                wb, ws_tpl,
                store=store,
                store_short_name=store_names.get(store, ""),
                dates=dates,
                day_map=topn_dict[store],
                cat_name=cat_name,
                event_name=event_name,
                total_all_dict=total_all_dict,
                total_cat_dict=total_cat_dict,
                category=category,
                make_title=make_title,
            )

        if "TEMPLATE" in wb.sheetnames:
            del wb["TEMPLATE"]

        wb.save(out_path)
        print(f"[ok] saved → {out_path}")


        # 店番スプリット（オプション）
        if split_by_store:
            save_per_store_files(Path(out_path), Path(split_dir), CATEGORY_MAP.get(str(category), str(category)))

# === TopN 作成（store×date×大分類で金額降順TopN） ===
def aggregate_topn(df_sales: pd.DataFrame, category: int, top_n: int = 35, dates=None):
    """
    df_sales : 列に date, store_id, category_large, jan, name, amount, (qty, discount 任意)
    dates    : list[date] or None
    戻り値   : dict[store_id -> dict[date -> DataFrame(TopN降順)]]
    """
    gdf = df_sales.copy()
    # 安全に型整形
    if "date" in gdf.columns:
        gdf["date"] = pd.to_datetime(gdf["date"], errors="coerce").dt.date
    if "store_id" in gdf.columns:
        gdf["store_id"] = gdf["store_id"].astype(str)
    if "category_large" in gdf.columns:
        gdf["category_large"] = gdf["category_large"].astype(str)

    # 日付フィルタ（指定時のみ）
    if dates:
        dt_set = set(pd.to_datetime(dates).date)
        gdf = gdf[gdf["date"].isin(dt_set)]

    # 大分類フィルタ
    gdf = gdf[gdf["category_large"] == str(category)]

    # 同一 (date, store, jan) を合算して 4倍問題を恒久対策
    agg_map = {"amount": "sum"}
    if "qty" in gdf.columns: agg_map["qty"] = "sum"
    if "discount" in gdf.columns: agg_map["discount"] = "sum"
    if "name" in gdf.columns: agg_map["name"] = "first"

    gdf = (
        gdf.groupby(["date", "store_id", "jan"], as_index=False, sort=False)
           .agg(agg_map)
    )

    # 金額降順でTopN抽出 → store×date の辞書に
    out = {}
    for (store, d), sub in gdf.groupby(["store_id", "date"], sort=False):
        sub = sub.sort_values("amount", ascending=False).head(top_n).reset_index(drop=True)
        out.setdefault(store, {})[d] = sub

    return out

if __name__ == "__main__":
    import argparse
    from pathlib import Path
    import pandas as pd

    parser = argparse.ArgumentParser(description="TopN distributor Excel generator (refactored)")
    parser.add_argument("--event-name", type=str, default="秋の感謝セール")
    parser.add_argument("--title-template",
                    type=str,
                    default="{yy}年 {range} {cat}単品データ（{page}）",
                    help=("A1タイトルのテンプレ。{event},{cat},{page} に加えて "
                          "{range},{dates},{dates_short},{category},{year},{yy},{date},{date_short} が利用可。"
                          "イベント名が空欄のとき自動で本テンプレが使用されます"))
    parser.add_argument("--no-date-in-title",
                        action="store_true",
                        help="タイトルから日付を除外（= '{event} {cat}単品データ ({page})'）")
    parser.add_argument("--category", type=int, required=True)
    parser.add_argument("--dates", type=str, required=True, help="YYYY-MM-DD をカンマ区切り")
    parser.add_argument("--out", type=str, required=True)
    parser.add_argument("--split-by-store", action="store_true",
                        help="店番ごとに別ファイルで出力する")
    parser.add_argument("--split-dir", type=str, default="",
                        help="店別ファイルの出力先ルート（未指定なら out と同階層に stores/）")
    args = parser.parse_args()

    print("[debug] 開始")
    proj_root = Path(__file__).resolve().parents[1]
    sales_root = proj_root / "data" / "material"
    template_path = proj_root / "data" / "template" / "配布フォーマット.xlsx"
    store_master = sales_root / "master" / "store_master.xlsx"

    dates = [pd.to_datetime(x).date() for x in args.dates.split(",")]
    df_sales = load_sales(sales_root, dates=dates)  # ← 年月またぎで必要なCSVだけ読む
    store_names = load_store_master(store_master)

    topn = aggregate_topn(df_sales, category=args.category, top_n=35, dates=dates)

    write_excel(
        template_path=template_path,
        out_path=Path(args.out),
        topn_dict=topn,
        store_names=store_names,
        category=args.category,
        dates=dates,
        event_name=args.event_name,
        df_sales_all=df_sales,         # 使っているなら
        split_by_store=args.split_by_store,
        split_dir=args.split_dir,      # ← これだけ渡す
        title_template=args.title_template,
        no_date_in_title=args.no_date_in_title,
    )


